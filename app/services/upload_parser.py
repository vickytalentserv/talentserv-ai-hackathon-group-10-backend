from __future__ import annotations

import csv
import io
from pathlib import Path

SUPPORTED_UPLOAD_EXTENSIONS = {".csv", ".xlsx"}
SUPPORTED_DATASET_TYPES = {"properties"}


def normalize_row(raw_row: dict[str, object | None]) -> dict[str, str]:
    return {
        str(key).strip(): "" if value is None else str(value).strip()
        for key, value in raw_row.items()
        if key is not None and str(key).strip()
    }


def parse_upload_rows(filename: str, content: bytes) -> list[dict[str, str]]:
    suffix = Path(filename).suffix.lower()
    if suffix not in SUPPORTED_UPLOAD_EXTENSIONS:
        supported = ", ".join(sorted(SUPPORTED_UPLOAD_EXTENSIONS))
        raise ValueError(f"Unsupported file type '{suffix or 'unknown'}'. Use: {supported}")

    if suffix == ".csv":
        return _parse_csv(content)

    return _parse_xlsx(content)


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        raise ValueError("CSV file is missing a header row")

    rows: list[dict[str, str]] = []
    for raw_row in reader:
        normalized = normalize_row(raw_row)
        if any(normalized.values()):
            rows.append(normalized)

    if not rows:
        raise ValueError("CSV file has no data rows")

    return rows


def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Excel support is not installed on the server") from exc

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active
    row_iter = worksheet.iter_rows(values_only=True)
    header_row = next(row_iter, None)

    if not header_row:
        raise ValueError("Excel file is empty")

    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    if not any(headers):
        raise ValueError("Excel file is missing a header row")

    rows: list[dict[str, str]] = []
    for row_values in row_iter:
        if row_values is None or all(value is None or str(value).strip() == "" for value in row_values):
            continue

        raw_row: dict[str, object | None] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            raw_row[header] = row_values[index] if index < len(row_values) else None

        normalized = normalize_row(raw_row)
        if any(normalized.values()):
            rows.append(normalized)

    if not rows:
        raise ValueError("Excel file has no data rows")

    return rows
